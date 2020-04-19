import discord
import openpyxl
import random

client = discord.Client()
admin_role = [651413680306651188, 642987212916391956, 651405902967996427, 642987402146873382, 661468035441360943]

@client.event
async def on_ready():
    print(client.user.id, "준비됐음다")
    print("https://discordapp.com/oauth2/authorize?client_id=655650874005127168permissions=8&scope=bot")

    game = discord.Game("학습")
    await client.change_presence(status=discord.Status.online, activity=game)


@client.event
async def on_message(message):
    if message.content.startswith("스쿨봇 ") or message.content.startswith("ㅅ "):
        if message.content.startswith("스쿨봇 도움") or message.content.startswith("ㅅ 도움"):
            embed = discord.Embed(
                title="**<스쿨봇의 도움말>**",
                description="TOP님을 위해 chang06과 자연화맛 음료수가 만든 봇이닷\n\n",
                colour=discord.Colour.green()
            )
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/651417391091417094/651726452714242059/a2e406895be038cb.png")
            embed.add_field(name='--------------------------__<관리자 전용 명령어>__-------------------------', value=
            "> 스쿨봇 올리기 [유저 맨션] (점수)\n"
            + "스쿨봇이 맨션한 유저의 지식점수를 [점수]만큼 증가 시킨다\n   "
            + "> 스쿨봇 내리기 [유저 맨션] (점수)\n"
            + "스쿨봇이 맨션한 유저의 지식점수를 (점수)만큼 감소 시킨다\n"
            + "> 스쿨봇 정하기 [유저 맨션] (점수)\n"
            + "스쿨봇이 맨션한 유저의 지식점수를 [점수]로 바꾼다\n", inline=False)
            embed.add_field(name='-----------------------------__<공용 명령어>__-----------------------------', value=
            "> 스쿨봇 지식점수 [유저 맨션]\n"
            + "스쿨봇이 맨션한 유저의 점수를 알려준다\n"
            + "> 스쿨봇 내 지식점수\n"
            + "스쿨봇이 명령자의 지식점수를 알려준다\n"
            + "> 스쿨봇 차트 \n"
            + "스쿨봇이 지식점수가 높은 유저들 부터 낮은 유저들 순으로 차트를 나열한다\n"
            + "> 스쿨봇 [아무말]\n"
            + "스쿨봇이 배운 말들 중 그에 맞는 대답을 한다\n"
            + "> 스쿨봇 말추가 [A]/[B]\n"
            + "> 스쿨봇 프로필 \n "
            + "디스코드 닉네임,별명,아이디를 볼수 있다."
            + "누군가 [A]라고 말하면 스쿨봇이 [B]라고 대답한다\n"
            + "예시 : `스쿨봇 말추가 안녕/ㅎㅇㅎㅇ`\n", inline=False)

            await message.channel.send(embed=embed)



        elif message.content.startswith("스쿨봇 올리기") or message.content.startswith("ㅅ 올리기"):
            if str(message.author.id) == "492222241501741056" or str(message.author.id) == "481077311890784276" or str(
                    message.author.id) == "550226629717131287" or str(message.author.id) == "542503495752876035":
                file = openpyxl.load_workbook("학생들.xlsx")
                sheet = file.active

                usaid = str(message.content)
                share = usaid.split()
                person = message.mentions[0].id

                i = 1
                flag = 0
                while sheet["A" + str(i)].value != None:
                    if str(sheet["A" + str(i)].value) == str(person):
                        sheet["B" + str(i)].value += int(share[3])
                        flag = 1

                        file.save("학생들.xlsx")
                        await message.channel.send(
                            "축하합니다! " + str(share[2]) + "님의 지식 점수가 " + str(share[3]) + "만큼 올랐습니다!\n(현재 : " + str(
                                sheet["B" + str(i)].value) + "점)")
                        break

                    i += 1

                if flag == 0:
                    sheet["A" + str(i)].value = str(person)
                    sheet["B" + str(i)].value = int(share[3])
                    sheet["C" + str(i)].value = str(message.mentions[0])

                    file.save("학생들.xlsx")
                    await message.channel.send(
                        "축하합니다! " + str(share[2]) + "님의 지식 점수가 " + str(share[3]) + "만큼 올랐습니다!\n(현재 : " + str(
                            sheet["B" + str(i)].value) + "점)")
            else:
                embed = discord.Embed(
                    title="**<Error! 삐비빅>**",
                    description="가소로운 뇨오석~\n이 명령어는 관리자만 쓸 수 있다굿!\n강해저서 돌아와라 닝겐(킹시국<<ㅍㅍ)",
                    colour=discord.Colour.red()
                )
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/651417391091417094/651726452714242059/a2e406895be038cb.png")
                await message.channel.send(embed=embed)

        elif message.content.startswith("스쿨봇 내리기") or message.content.startswith("ㅅ 내리기"):
            if str(message.author.id) == "492222241501741056" or str(message.author.id) == "481077311890784276" or str(
                    message.author.id) == "550226629717131287" or str(message.author.id) == "542503495752876035":
                file = openpyxl.load_workbook("학생들.xlsx")
                sheet = file.active

                usaid = str(message.content)
                share = usaid.split()
                person = message.mentions[0].id

                i = 1
                flag = 0
                while sheet["A" + str(i)].value != None:
                    if str(sheet["A" + str(i)].value) == str(person):
                        sheet["B" + str(i)].value -= int(share[3])
                        flag = 1

                        file.save("학생들.xlsx")
                        await message.channel.send(
                            str(share[2]) + "님의 지식 점수가 " + str(share[3]) + "만큼 줄었습니다...\n(현재 : " + str(
                                sheet["B" + str(i)].value) + "점)")
                        break

                    i += 1

                if flag == 0:
                    await message.channel.send(":thinking:   그런 학생을 찾을 수가 없네요")
            else:
                embed = discord.Embed(
                    title="**<Error! 삐비빅>**",
                    description="가소로운 뇨오석~\n이 명령어는 관리자만 쓸 수 있다굿!\n강해저서 돌아와라 닝겐(킹시국<<ㅍㅍ)",
                    colour=discord.Colour.red()
                )
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/651417391091417094/651726452714242059/a2e406895be038cb.png")
                await message.channel.send(embed=embed)

        elif message.content.startswith("스쿨봇 정하기") or message.content.startswith("ㅅ 정하기"):
            if str(message.author.id) == "492222241501741056" or str(message.author.id) == "481077311890784276" or str(
                    message.author.id) == "550226629717131287" or str(message.author.id) == "542503495752876035":
                file = openpyxl.load_workbook("학생들.xlsx")
                sheet = file.active

                usaid = str(message.content)
                share = usaid.split()
                person = message.mentions[0].id

                i = 1
                flag = 0
                while sheet["A" + str(i)].value != None:
                    if str(sheet["A" + str(i)].value) == str(person):
                        sheet["B" + str(i)].value = int(share[3])
                        flag = 1

                        file.save("학생들.xlsx")
                        await message.channel.send(str(share[2]) + "님의 지식 점수가 " + str(share[3]) + "로 바뀌었습니닷")
                        break

                    i += 1

                if flag == 0:
                    sheet["A" + str(i)].value = str(person)
                    sheet["B" + str(i)].value = int(share[3])
                    sheet["C" + str(i)].value = str(message.mentions[0])

                    file.save("학생들.xlsx")
                    await message.channel.send(str(share[2]) + "님의 지식 점수가 " + str(share[3]) + "로 바뀌었습니닷")
            else:
                embed = discord.Embed(
                    title="**<Error! 삐비빅>**",
                    description="가소로운 뇨오석~\n이 명령어는 관리자만 쓸 수 있다굿!\n강해저서 돌아와라 닝겐(킹시국<<ㅍㅍ)",
                    colour=discord.Colour.red()
                )
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/651417391091417094/651726452714242059/a2e406895be038cb.png")
                await message.channel.send(embed=embed)

        elif message.content.startswith("스쿨봇 지식점수") or message.content.startswith("ㅅ 지식점수"):
            if message.mentions[0].id == 651412998191185941:
                await message.channel.send("몇 랩인지 궁금하시나요?\n||저는 만랩이라 999999999999999999 99999999랩입니다!ㅋㅋ||")
            else:
                file = openpyxl.load_workbook("학생들.xlsx")
                sheet = file.active

                usaid = str(message.content)
                share = usaid.split()
                mention = share[2]
                person = message.mentions[0].id

                i = 1
                flag = 0
                while sheet["A" + str(i)].value != None:
                    if str(sheet["A" + str(i)].value) == str(person):
                        flag = 1

                        file.save("학생들.xlsx")
                        await message.channel.send(
                            str(mention) + "님의 지식 점수는 바로 " + str(sheet["B" + str(i)].value) + "점입니닷")
                        break

                    i += 1

                if flag == 0:
                    embed = discord.Embed(
                        title="**<Error! 삐비빅>**",
                        description="그런 사람은 잘 모르겠네요\n`스쿨봇 올리기`를 먼저 사용해주세욥",
                        colour=discord.Colour.red()
                    )
                    embed.set_thumbnail(
                        url="https://cdn.discordapp.com/attachments/651417391091417094/651726452714242059/a2e406895be038cb.png")
                    await message.channel.send(embed=embed)

        elif message.content == "스쿨봇 내 지식점수" or message.content == "ㅅ 내 지식점수":
            file = openpyxl.load_workbook("학생들.xlsx")
            sheet = file.active

            usaid = str(message.content)
            share = usaid.split()

            i = 1
            flag = 0
            while sheet["A" + str(i)].value != None:
                if str(sheet["A" + str(i)].value) == str(message.author.id):
                    flag = 1

                    file.save("학생들.xlsx")
                    await message.channel.send("당신의 지식 점수는 바로 " + str(sheet["B" + str(i)].value) + "점입니닷")
                    break

                i += 1

            if flag == 0:
                embed = discord.Embed(
                    title="**<Error! 삐비빅>**",
                    description="그런 사람은 잘 모르겠네요\n`스쿨봇 올리기`를 먼저 사용해주세욥",
                    colour=discord.Colour.red()
                )
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/651417391091417094/651726452714242059/a2e406895be038cb.png")
                await message.channel.send(embed=embed)

        elif message.content == "스쿨봇 차트" or message.content == "ㅅ 차트":
            file = openpyxl.load_workbook("학생들.xlsx")
            sheet = file.active

            i = 1
            while sheet["A" + str(i)].value != None:
                i += 1

            arr = [0] * i
            user = []
            num = []
            real_num = 1
            number = 1
            ln = -2147483647
            ans = ""
            while arr.count(0) > 1:
                i = 1
                m = -2147483648
                mi = 0
                while sheet["A" + str(i)].value != None:
                    if arr[i] == 0 and sheet["B" + str(i)].value > m:
                        m = sheet["B" + str(i)].value
                        mi = i
                    i += 1

                if int(sheet["B" + str(mi)].value) == 0: break

                arr[mi] = 1
                real_num += 1
                if sheet["B" + str(mi)].value != ln: number = real_num

                if sheet["B" + str(mi)].value != ln: ans = ans + "[" + str(number-1) + "위] : " + sheet["C" + str(mi)].value[:(len(sheet["C" + str(mi)].value) - 5)] + "  <" + str(sheet["B" + str(mi)].value) + ">\n"
                else: ans = ans + "[공동   " + str(number-1) + "위] : " + sheet["C" + str(mi)].value[:(len(sheet["C" + str(mi)].value) - 5)] + "  <" + str(sheet["B" + str(mi)].value) + ">\n"

                ln = sheet["B" + str(mi)].value

            embed = discord.Embed(title="**<지식 점수 순위>**", description=ans, colour=discord.Colour.gold())
            embed.set_thumbnail(url = "https://us.123rf.com/450wm/oakozhan/oakozhan1711/oakozhan171100015/89473349-3-%EC%88%9C%EC%9C%84-%EC%9E%A5%EC%86%8C-3-%EC%B0%A8%EC%9B%90-%EB%A0%8C%EB%8D%94%EB%A7%81-3d-%EC%9D%BC%EB%9F%AC%EC%8A%A4%ED%8A%B8%EC%99%80-%ED%9D%B0%EC%83%89-%EC%8B%A4%EB%A6%B0%EB%8D%94-%EC%97%B0%EB%8B%A8.jpg?ver=6")
            await message.channel.send(embed=embed)

        elif message.content.startswith("스쿨봇 말추가") or message.content.startswith("ㅅ 말추가"):
            file = openpyxl.load_workbook("배운말.xlsx")
            sheet = file.active

            q = message.content.split("/")[0]
            a = message.content.split("/")[1]

            if message.content.startswith("스쿨봇 "):
                q2 = q[8:]
            elif message.content.startswith("ㅅ "):
                q2 = q[6:]

            i = 1
            flag = 0
            while sheet["A" + str(i)].value != None:
                if sheet["A" + str(i)].value == q2: flag = 1
                i += 1

            if flag == 0:
                sheet["A" + str(i)].value = str(q2)
                sheet["B" + str(i)].value = str(a)
                sheet["C" + str(i)].value = str(message.author.id)
                sheet["D" + str(i)].value = str(message.author)

                await message.channel.send("[" + str(q2) + "]라고 하면 [" + str(a) + "]라고 대답하는 것을 배웠어!")
                file.save("배운말.xlsx")
            else:
                embed = discord.Embed(
                    title="**<Error! 삐비빅>**",
                    description="가소로운 뇨오석~\n그건 이미 배운 말이라굿!\n강해저서 돌아와라 닝겐(킹시국<<ㅍㅍ)",
                    colour=discord.Colour.red()
                )
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/651417391091417094/651726452714242059/a2e406895be038cb.png")
                await message.channel.send(embed=embed)

        elif message.content.startswith("스쿨봇 미션추가") or message.content.startswith("ㅅ 미션추가"):
            if str(message.author.id) == "492222241501741056" or str(message.author.id) == "481077311890784276" or str(
                    message.author.id) == "550226629717131287" or str(message.author.id) == "542503495752876035":
                file = openpyxl.load_workbook("미션.xlsx")
                sheet = file.active

                if message.content.startswith("스쿨봇 미션추가"): mission = message.content[9:]
                elif message.content.startswith("ㅅ 미션추가"): mission = message.content[7:]

                i = 1
                while sheet["A" + str(i)].value != None:
                    i += 1

                sheet["A" + str(i)].value = str(mission)

                await message.channel.send("[" + str(mission) + "]라는 미션을 배웠어!")
                file.save("미션.xlsx")
            else:
                embed = discord.Embed(
                    title="**<Error! 삐비빅>**",
                    description="가소로운 뇨오석~\n이 명령어는 관리자만 쓸 수 있다굿!\n강해저서 돌아와라 닝겐(킹시국<<ㅍㅍ)",
                    colour=discord.Colour.red()
                )
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/651417391091417094/651726452714242059/a2e406895be038cb.png")
                await message.channel.send(embed=embed)

        elif message.content == "스쿨봇 퀘스트" or message.content == "ㅅ 퀘스트":
            file = openpyxl.load_workbook("미션.xlsx")
            sheet = file.active

            i, mission = 1, ""
            while sheet["A" + str(i)].value != None:
                mission = mission + str(sheet["A" + str(i)].value) + "\n"
                i += 1

            if i > 1: embed = discord.Embed(title="퀘스트 목록", description=mission, colour=discord.Colour.green())
            else : embed = discord.Embed(title="퀘스트 목록", description="아직 퀘스트가 없네요... ", colour=discord.Colour.green())
            await message.channel.send(embed=embed)

        elif message.content.startswith("스쿨봇 미션클리어") or message.content.startswith("ㅅ 미션클리어"):
            if str(message.author.id) == "492222241501741056" or str(message.author.id) == "481077311890784276" or str(
                    message.author.id) == "550226629717131287" or str(message.author.id) == "542503495752876035":
                m_file = openpyxl.load_workbook("미션.xlsx")
                m_sheet = m_file.active

                c_file = openpyxl.load_workbook("미션클리어.xlsx")
                c_sheet = c_file.active

                person = message.mentions[0].id

                space = 0
                for i in range(len(message.content)):
                    if message.content[i] == ' ': space += 1
                    if space == 3: break
                    i += 1

                mission = message.content[i+1:]

                mn, i = -1, 1
                while m_sheet["A" + str(i)].value != None:
                    if m_sheet["A" + str(i)].value == mission:
                        mn = i
                        break
                    i += 1

                if mn != -1:
                    i, flag = 1, 0
                    while c_sheet["A" + str(i)].value != None:
                        if c_sheet["A" + str(i)].value == str(person) and c_sheet["B" + str(i)].value == str(mn):
                            flag = 1
                        i += 1

                    if flag == 0:
                        i = 1
                        while c_sheet["A" + str(i)].value != None:
                            i += 1

                        c_sheet["A" + str(i)].value = str(person)
                        c_sheet["B" + str(i)].value = str(mn)

                        c_file.save("미션클리어.xlsx")
                        await message.channel.send("미션을 클리어하신걸 축하드립니다!")

                    else : await message.channel.send("이 사람은 이미 그 미션을 깼던 거 같아")
                else : await message.channel.send("그건 없는 미션 같아")
            else : await message.channel.send("이건 권한 있어야 ㄱㄴ")

        elif message.content.startswith("ㅅ 공지"):
            if str(message.author.id) == "492222241501741056" or str(message.author.id) == "481077311890784276" or str(
                    message.author.id) == "550226629717131287" or str(message.author.id) == "542503495752876035" or str(message.author.id) == "477280468400865280":

                channel = message.content[5:23]
                msg = message.content[24:]
                await client.get_channel(int(channel)).send(msg)
            else:
                embed = discord.Embed(
                    title="**<Error! 삐비빅>**",
                    description="가소로운 뇨오석~\n이 명령어는 관리자만 쓸 수 있다굿!\n강해저서 돌아와라 닝겐(킹시국<<ㅍㅍ)",
                    colour=discord.Colour.red()
                )
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/651417391091417094/651726452714242059/a2e406895be038cb.png")
                await message.channel.send(embed=embed)





        elif message.content.startswith("스쿨봇 공지"):
            if str(message.author.id) == "492222241501741056" or str(message.author.id) == "481077311890784276" or str(
                    message.author.id) == "550226629717131287" or str(message.author.id) == "542503495752876035" or str(message.author.id) == "477280468400865280":

                channel = message.content[5:23]
                msg = message.content[24:]
                await client.get_channel(int(channel)).send(msg)
            else:
                embed = discord.Embed(
                    title="**<Error! 삐비빅>**",
                    description="가소로운 뇨오석~\n이 명령어는 관리자만 쓸 수 있다굿!\n강해저서 돌아와라 닝겐(킹시국<<ㅍㅍ)",
                    colour=discord.Colour.red()
                )
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/651417391091417094/651726452714242059/a2e406895be038cb.png")
                await message.channel.send(embed=embed)



        elif message.content.startswith("스쿨봇 프로필"):
            file = openpyxl.load_workbook("학생들.xlsx")
            sheet = file.active

            usaid = str(message.content)
            share = usaid.split()

            i = 1
            flag = 0
            while sheet["A" + str(i)].value != None:
                if str(sheet["A" + str(i)].value) == str(message.author.id):
                    flag = 1

                    file.save("학생들.xlsx")
                    break

                i += 1

            if flag == 0:
                embed = discord.Embed(
                    title="**<Error! 삐비빅>**",
                    description="저런... 지식점수가 없어요.\n`스쿨봇 올리기`를 먼저 사용해주세욥",
                    colour=discord.Colour.red()
                )
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/651417391091417094/651726452714242059/a2e406895be038cb.png")
                await message.channel.send(embed=embed)

            embed = discord.Embed(color=0x00ff00)
            embed.add_field(name="이름", value=message.author.name, inline=True)
            embed.add_field(name="서버닉네임", value=message.author.display_name, inline=True)
            embed.add_field(name="아이디", value=message.author.id, inline=True)
            embed.add_field(name="지식점수", value=str(sheet["B" + str(i)].value) + "점", inline=True)
            embed.set_thumbnail(url=message.author.avatar_url)
            await message.channel.send(embed=embed)


        elif message.content.startswith("ㅅ 프로필"):
            file = openpyxl.load_workbook("학생들.xlsx")
            sheet = file.active

            usaid = str(message.content)
            share = usaid.split()

            i = 1
            flag = 0
            while sheet["A" + str(i)].value != None:
                if str(sheet["A" + str(i)].value) == str(message.author.id):
                    flag = 1

                    file.save("학생들.xlsx")
                    break

                i += 1

            if flag == 0:
                embed = discord.Embed(
                    title="**<Error! 삐비빅>**",
                    description="저런... 지식점수가 없어요.\n 열심히 활동해서 지식을 얻으세요",
                    colour=discord.Colour.red()
                )
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/651417391091417094/651726452714242059/a2e406895be038cb.png")
                await message.channel.send(embed=embed)
            embed = discord.Embed(color=0x00ff00)
            embed.add_field(name="이름", value=message.author.name, inline=True)
            embed.add_field(name="서버닉네임", value=message.author.display_name, inline=True)
            embed.add_field(name="아이디", value=message.author.id, inline=True)
            embed.add_field(name="지식점수", value=str(sheet["B" + str(i)].value) + "점", inline=True)
            embed.set_thumbnail(url=message.author.avatar_url)
            await message.channel.send(embed=embed)

        elif message.content.startswith("스쿨봇 명예의전당"):
            file = openpyxl.load_workbook("명예의전당.xlsx")
            sheet = file.active

            i, mission = 1, ""
            while sheet["A" + str(i)].value != None:
                mission = mission + str(sheet["A" + str(i)].value) + "\n"
                i += 1

            if i > 1:
                embed = discord.Embed(title="**<지식 명예의전당 >**", description=mission, colour=discord.Colour.gold())
                embed.set_thumbnail(
                    url="https://us.123rf.com/450wm/oakozhan/oakozhan1711/oakozhan171100015/89473349-3-%EC%88%9C%EC%9C%84-%EC%9E%A5%EC%86%8C-3-%EC%B0%A8%EC%9B%90-%EB%A0%8C%EB%8D%94%EB%A7%81-3d-%EC%9D%BC%EB%9F%AC%EC%8A%A4%ED%8A%B8%EC%99%80-%ED%9D%B0%EC%83%89-%EC%8B%A4%EB%A6%B0%EB%8D%94-%EC%97%B0%EB%8B%A8.jpg?ver=6")
            else:
                embed = discord.Embed(title="지식 명예의전당", description="아직 기록이 없네요... ", colour=discord.Colour.green())
            await message.channel.send(embed=embed)




        elif message.content.startswith("ㅅ 명예의전당"):
            file = openpyxl.load_workbook("명예의전당.xlsx")
            sheet = file.active

            i, mission = 1, ""
            while sheet["A" + str(i)].value != None:
                mission = mission + str(sheet["A" + str(i)].value) + "\n"
                i += 1

            if i > 1:
                embed = discord.Embed(title="**<지식 명예의전당 >**", description=mission, colour=discord.Colour.gold())
                embed.set_thumbnail(
                    url="https://us.123rf.com/450wm/oakozhan/oakozhan1711/oakozhan171100015/89473349-3-%EC%88%9C%EC%9C%84-%EC%9E%A5%EC%86%8C-3-%EC%B0%A8%EC%9B%90-%EB%A0%8C%EB%8D%94%EB%A7%81-3d-%EC%9D%BC%EB%9F%AC%EC%8A%A4%ED%8A%B8%EC%99%80-%ED%9D%B0%EC%83%89-%EC%8B%A4%EB%A6%B0%EB%8D%94-%EC%97%B0%EB%8B%A8.jpg?ver=6")
            else:
                embed = discord.Embed(title="지식 명예의전당", description="아직 기록이 없네요... ", colour=discord.Colour.green())
            await message.channel.send(embed=embed)


        elif message.content.startswith("스쿨봇 거래소"):
            file = openpyxl.load_workbook("거래소.xlsx")
            sheet = file.active

            i, mission = 1, ""
            while sheet["A" + str(i)].value != None:
                mission = mission + str(sheet["A" + str(i)].value) + "\n"
                i += 1

            if i > 1:
                embed = discord.Embed(title="**<지식 거래소 >**", description=mission, colour=discord.Colour.gold())
            else:
                embed = discord.Embed(title="지식 거래소", description="아직 물품이 없네요... ", colour=discord.Colour.green())
            await message.channel.send(embed=embed)

        elif message.content.startswith("ㅅ 거래소"):
            file = openpyxl.load_workbook("거래소.xlsx")
            sheet = file.active

            i, mission = 1, ""
            while sheet["A" + str(i)].value != None:
                mission = mission + str(sheet["A" + str(i)].value) + "\n"
                i += 1

            if i > 1:
                embed = discord.Embed(title="**<지식 거래소 >**", description=mission, colour=discord.Colour.gold())
            else:
                embed = discord.Embed(title="지식 거래소", description="아직 물품이 없네요... ", colour=discord.Colour.green())
            await message.channel.send(embed=embed)




        elif message.content.startswith("스쿨봇 데스노트 올리기") or message.content.startswith("ㅅ 데스노트 올리기"):
            if str(message.author.id) == "492222241501741056" or str(message.author.id) == "481077311890784276" or str(
                    message.author.id) == "550226629717131287" or str(message.author.id) == "542503495752876035":
                file = openpyxl.load_workbook("데스노트 랭킹.xlsx")
                sheet = file.active

                usaid = str(message.content)
                share = usaid.split()
                person = message.mentions[0].id

                i = 1
                flag = 0
                while sheet["A" + str(i)].value != None:
                    if str(sheet["A" + str(i)].value) == str(person):
                        sheet["B" + str(i)].value += int(share[3])
                        flag = 1

                        file.save("데스노트 랭킹.xlsx")
                        await message.channel.send(
                            "축하합니다! " + str(share[2]) + "님의 RP가 " + str(share[3]) + "만큼 올랐습니다!\n(현재 : " + str(
                                sheet["B" + str(i)].value) + "점)")
                        break

                    i += 1

                if flag == 0:
                    sheet["A" + str(i)].value = str(person)
                    sheet["B" + str(i)].value = int(share[3])
                    sheet["C" + str(i)].value = str(message.mentions[0])

                    file.save("데스노트 랭킹.xlsx")
                    await message.channel.send(
                        "축하합니다! " + str(share[2]) + "님의 RP가 " + str(share[3]) + "만큼 올랐습니다!\n(현재 : " + str(
                            sheet["B" + str(i)].value) + "점)")
            else:
                embed = discord.Embed(
                    title="**<Error! 삐비빅>**",
                    description="가소로운 뇨오석~\n이 명령어는 관리자만 쓸 수 있다굿!\n강해저서 돌아와라 닝겐(킹시국<<ㅍㅍ)",
                    colour=discord.Colour.red()
                )
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/651417391091417094/651726452714242059/a2e406895be038cb.png")
                await message.channel.send(embed=embed)



        elif message.content.startswith("스쿨봇 데스노트 내리기") or message.content.startswith("ㅅ 데스노트 내리기"):
                    if str(message.author.id) == "492222241501741056" or str(
                            message.author.id) == "481077311890784276" or str(
                            message.author.id) == "550226629717131287" or str(
                        message.author.id) == "542503495752876035":
                        file = openpyxl.load_workbook("데스노트 랭킹.xlsx")
                        sheet = file.active

                        usaid = str(message.content)
                        share = usaid.split()
                        person = message.mentions[0].id

                        i = 1
                        flag = 0
                        while sheet["A" + str(i)].value != None:
                            if str(sheet["A" + str(i)].value) == str(person):
                                sheet["B" + str(i)].value -= int(share[3])
                                flag = 1

                                file.save("데스노트 랭킹.xlsx")
                                await message.channel.send(
                                    str(share[2]) + "님의 RP가 " + str(share[3]) + "만큼 줄었습니다...\n(현재 : " + str(
                                        sheet["B" + str(i)].value) + "점)")
                                break

                            i += 1

                        if flag == 0:
                            await message.channel.send(":thinking:   그런 학생을 찾을 수가 없네요")
                    else:
                        embed = discord.Embed(
                            title="**<Error! 삐비빅>**",
                            description="가소로운 뇨오석~\n이 명령어는 관리자만 쓸 수 있다굿!\n강해저서 돌아와라 닝겐(킹시국<<ㅍㅍ)",
                            colour=discord.Colour.red()
                        )
                        embed.set_thumbnail(
                            url="https://cdn.discordapp.com/attachments/651417391091417094/651726452714242059/a2e406895be038cb.png")
                        await message.channel.send(embed=embed)

        elif message.content == "스쿨봇 랭킹" or message.content =="ㅅ 랭킹":
            file = openpyxl.load_workbook("데스노트 랭킹.xlsx")
            sheet = file.active

            i = 1
            while sheet["A" + str(i)].value != None:
                i += 1

            arr = [0] * i
            user = []
            num = []
            real_num = 1
            number = 1
            ln = -2147483647
            ans = ""
            while arr.count(0) > 1:
                i = 1
                m = -2147483648
                mi = 0
                while sheet["A" + str(i)].value != None:
                    if arr[i] == 0 and sheet["B" + str(i)].value > m:
                        m = sheet["B" + str(i)].value
                        mi = i
                    i += 1

                if int(sheet["B" + str(mi)].value) == 0: break

                arr[mi] = 1
                real_num += 1
                if sheet["B" + str(mi)].value != ln: number = real_num

                if sheet["B" + str(mi)].value != ln: ans = ans + "[" + str(number-1) + "위] : " + sheet["C" + str(mi)].value[:(len(sheet["C" + str(mi)].value) - 5)] + "  <" + str(sheet["B" + str(mi)].value) + ">\n"
                else: ans = ans + "[공동   " + str(number-1) + "위] : " + sheet["C" + str(mi)].value[:(len(sheet["C" + str(mi)].value) - 5)] + "  <" + str(sheet["B" + str(mi)].value) + ">\n"

                ln = sheet["B" + str(mi)].value

            embed = discord.Embed(title="**<데스노트 RP 순위>**", description=ans, colour=discord.Colour.gold())
            embed.set_thumbnail(url = "https://us.123rf.com/450wm/oakozhan/oakozhan1711/oakozhan171100015/89473349-3-%EC%88%9C%EC%9C%84-%EC%9E%A5%EC%86%8C-3-%EC%B0%A8%EC%9B%90-%EB%A0%8C%EB%8D%94%EB%A7%81-3d-%EC%9D%BC%EB%9F%AC%EC%8A%A4%ED%8A%B8%EC%99%80-%ED%9D%B0%EC%83%89-%EC%8B%A4%EB%A6%B0%EB%8D%94-%EC%97%B0%EB%8B%A8.jpg?ver=6")
            await message.channel.send(embed=embed)

        elif message.content.startswith("스쿨봇 데스노트 정하기") or message.content.startswith("ㅅ 데스노트 정하기"):
            if str(message.author.id) == "492222241501741056" or str(message.author.id) == "481077311890784276" or str(
                    message.author.id) == "550226629717131287" or str(message.author.id) == "542503495752876035":
                file = openpyxl.load_workbook("데스노트 랭킹.xlsx")
                sheet = file.active

                usaid = str(message.content)
                share = usaid.split()
                person = message.mentions[0].id

                i = 1
                flag = 0
                while sheet["A" + str(i)].value != None:
                    if str(sheet["A" + str(i)].value) == str(person):
                        sheet["B" + str(i)].value = int(share[3])
                        flag = 1

                        file.save("데스노트 랭킹.xlsx")
                        await message.channel.send(str(share[2]) + "님의 RP가 " + str(share[3]) + "로 바뀌었습니닷")
                        break

                    i += 1

                if flag == 0:
                    sheet["A" + str(i)].value = str(person)
                    sheet["B" + str(i)].value = str(share[3])
                    sheet["C" + str(i)].value = str(message.mentions[0])

                    file.save("데스노트 랭킹.xlsx")
                    await message.channel.send(str(share[3]) + "님의 RP가 " + str(share[4]) + "로 바뀌었습니닷")
            else:
                embed = discord.Embed(
                    title="**<Error! 삐비빅>**",
                    description="가소로운 뇨오석~\n이 명령어는 관리자만 쓸 수 있다굿!\n강해저서 돌아와라 닝겐(킹시국<<ㅍㅍ)",
                    colour=discord.Colour.red()
                )
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/651417391091417094/651726452714242059/a2e406895be038cb.png")
                await message.channel.send(embed=embed)


        elif message.content.startswith("ㅅ 역할지급"):
            author = message.guild.get_member(int(message.author.id))
            role = discord.utils.get(message.guild.roles, name="")
            await author.add_roles(role)










        else:
            file = openpyxl.load_workbook("배운말.xlsx")
            sheet = file.active

            if message.content.startswith("스쿨봇 "):
                q = message.content[4:]
            elif message.content.startswith("ㅅ "):
                q = message.content[2:]

            i = 1
            flag = 0
            while sheet["A" + str(i)].value != None:
                if sheet["A" + str(i)].value == q:
                    await message.channel.send(str(sheet["B" + str(i)].value))
                    flag = 1

                i += 1

            if flag == 0: await message.channel.send("님이 뭐라고 하시는지 잘 모루게쒀요\n`스쿨봇 도움`을 입력하셔서 명령어를 확인해주세욥")


client.run("NjYwODAyNzUwOTM4MDg3NDI2.XhqxMw.DwSTyNjrpU3WM3Sji3qYtQdFPgI")                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                              